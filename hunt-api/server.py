import os
import time
import json
import uuid
from io import BytesIO
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
CORS(app)

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {
        "users": {},
        "leaderboard_free": [],
        "leaderboard_paid": [],
        "giftcards": {},
        "user_giftcards": {},
        "winnings_last_checked": {},
    }


def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


# ---------------------------------------------------------------------------
# Leaderboard endpoints
# ---------------------------------------------------------------------------

@app.route("/api/leaderboard/free", methods=["GET"])
def get_free_leaderboard():
    data = load_data()
    entries = sorted(data.get("leaderboard_free", []), key=lambda e: e.get("score", 0), reverse=True)
    for idx, entry in enumerate(entries):
        entry["placement"] = idx + 1
    return jsonify(entries)


@app.route("/api/leaderboard/paid", methods=["GET"])
def get_paid_leaderboard():
    data = load_data()
    entries = sorted(data.get("leaderboard_paid", []), key=lambda e: e.get("score", 0), reverse=True)
    for idx, entry in enumerate(entries):
        entry["placement"] = idx + 1
    return jsonify(entries)


@app.route("/api/leaderboard", methods=["POST"])
def update_leaderboard_entry():
    body = request.get_json(force=True)
    username = body.get("username")
    email = body.get("email", "")
    score = body.get("score", 0)
    board_type = body.get("type", "free")

    if not username:
        return jsonify({"error": "username is required"}), 400
    if board_type not in ("free", "paid"):
        return jsonify({"error": "type must be 'free' or 'paid'"}), 400

    data = load_data()
    key = f"leaderboard_{board_type}"
    board = data.get(key, [])

    existing = next((e for e in board if e["username"] == username), None)
    if existing:
        existing["score"] = score
        if email:
            existing["email"] = email
    else:
        board.append({"username": username, "email": email, "score": score})

    data[key] = board
    save_data(data)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Giftcard endpoints
# ---------------------------------------------------------------------------

@app.route("/api/giftcard", methods=["POST"])
def create_giftcard():
    body = request.get_json(force=True)
    code = body.get("code")
    value = body.get("value")
    card_type = body.get("type", "")

    if not code or value is None:
        return jsonify({"error": "code and value are required"}), 400

    data = load_data()
    giftcards = data.get("giftcards", {})
    giftcards[code] = {
        "code": code,
        "value": value,
        "type": card_type,
        "assigned_to": None,
        "created_at": time.time(),
    }
    data["giftcards"] = giftcards
    save_data(data)
    return jsonify({"success": True, "giftcard": giftcards[code]})


@app.route("/api/giftcard/assign", methods=["POST"])
def assign_giftcard():
    body = request.get_json(force=True)
    code = body.get("code")
    username = body.get("username")

    if not code or not username:
        return jsonify({"error": "code and username are required"}), 400

    data = load_data()
    giftcards = data.get("giftcards", {})
    if code not in giftcards:
        return jsonify({"error": "giftcard not found"}), 404

    giftcards[code]["assigned_to"] = username
    giftcards[code]["assigned_at"] = time.time()

    user_gc = data.get("user_giftcards", {})
    if username not in user_gc:
        user_gc[username] = []
    user_gc[username].append({
        "code": code,
        "value": giftcards[code]["value"],
        "type": giftcards[code]["type"],
        "assigned_at": giftcards[code]["assigned_at"],
    })
    data["user_giftcards"] = user_gc
    data["giftcards"] = giftcards
    save_data(data)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Winnings endpoint – returns giftcards received since last check
# ---------------------------------------------------------------------------

@app.route("/api/winnings", methods=["GET"])
def get_winnings():
    username = request.args.get("username")
    if not username:
        return jsonify({"error": "username query parameter is required"}), 400

    data = load_data()
    last_checked = data.get("winnings_last_checked", {}).get(username, 0)

    user_gc = data.get("user_giftcards", {}).get(username, [])
    new_winnings = [gc for gc in user_gc if gc.get("assigned_at", 0) > last_checked]

    # Update last-checked timestamp
    if "winnings_last_checked" not in data:
        data["winnings_last_checked"] = {}
    data["winnings_last_checked"][username] = time.time()
    save_data(data)

    if not new_winnings:
        return jsonify({"message": "No new winnings.", "winnings": []})

    return jsonify({
        "message": f"You have {len(new_winnings)} new giftcard(s)!",
        "winnings": new_winnings,
    })


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@app.route("/api/export", methods=["GET"])
def export_excel():
    data = load_data()

    wb = Workbook()

    # ---- Free leaderboard sheet ----
    ws_free = wb.active
    ws_free.title = "Free Leaderboard"
    _fill_leaderboard_sheet(ws_free, data.get("leaderboard_free", []))

    # ---- Paid leaderboard sheet ----
    ws_paid = wb.create_sheet("Paid Leaderboard")
    _fill_leaderboard_sheet(ws_paid, data.get("leaderboard_paid", []))

    # ---- Giftcards sheet ----
    ws_gc = wb.create_sheet("Giftcards")
    _fill_giftcards_sheet(ws_gc, data.get("giftcards", {}))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="hunt_data.xlsx",
    )


def _fill_leaderboard_sheet(ws, entries):
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Placement", "Username", "Email", "Score"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    sorted_entries = sorted(entries, key=lambda e: e.get("score", 0), reverse=True)

    for row_idx, entry in enumerate(sorted_entries, start=2):
        placement = row_idx - 1
        ws.cell(row=row_idx, column=1, value=placement).border = thin_border
        ws.cell(row=row_idx, column=2, value=entry.get("username", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=entry.get("email", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=entry.get("score", 0)).border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20


def _fill_giftcards_sheet(ws, giftcards):
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Code", "Value", "Type", "Assigned To"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, gc in enumerate(giftcards.values(), start=2):
        ws.cell(row=row_idx, column=1, value=gc.get("code", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=gc.get("value", 0)).border = thin_border
        ws.cell(row=row_idx, column=3, value=gc.get("type", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=gc.get("assigned_to", "")).border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
